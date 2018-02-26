import urllib.request
import bs4 as bs
import datetime
import openpyxl
import difflib
import smtplib
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
import pynotify , os ,time
from flask import Flask, render_template, flash, request
from wtforms import Form, TextField, TextAreaField, validators, StringField, SubmitField

# App config.
DEBUG = True
app = Flask(__name__)
app.config.from_object(__name__)
app.config['SECRET_KEY'] = '7d441f27d441f27567d441f2b6176a'


class ReusableForm(Form):
    name = TextField('İsim:', validators=[validators.required()])
    mail = TextField("Email: ", validators=[validators.required()])


dataAdi = data = "title"

def calistir (ad, mail):
    while (1):
        for kaynak in range(1, 5):

            print(kaynak)
            print(type(kaynak))
            # webLink = 'https://ubersem.com/'
            # sendMail = 'tolga_k94@hotmail.com'

            webLink = ad
            sendMail = mail
            sauce = urllib.request.urlopen(webLink.strip()).read()
            soup = bs.BeautifulSoup(sauce, 'lxml')

            try:
                titledata = soup.title
                titletext = titledata.text

                h1data = soup.h1
                h1text = h1data.text

                h2data = soup.h2
                h2text = h2data.text
                
                bodydata = soup.body
                bodytext = bodydata.text

            except:
                print("Bu URL, girilen datalara ulaşılmasına izin vermiyor.")
            global dataAdi, data
            if request.form.get('boxtitle'):
                if (kaynak == 1):
                    data = titletext
                    dataAdi = "title"
            if request.form.get('boxh1'):
                if (kaynak == 2):
                    data = h1text
                    dataAdi = "h1"
            if request.form.get('boxh2'):
                if (kaynak == 3):
                    data = h2text
                    dataAdi = "h2"
            if request.form.get('boxbody'):
                if (kaynak == 4):
                    data = bodytext
                    dataAdi = "body"


            file = str(dataAdi + "DataBase.xlsx")
            print("KAYNAK NUMARASI: " + str(kaynak))

            def degisiklik(Ci_Verisi):
                if (Ci_Verisi == data):
                    print('Değişiklik yoktur')

                else:
                    textPD = "Veri değişikliği algılandı"
                    sayfa.append({'C': data, 'A': datetime.datetime.now(), 'I': webLink})

                    text1_lines = data.splitlines()
                    text2_lines = Ci_Verisi.splitlines()

                    d = difflib.Differ()
                    diff = d.compare(text1_lines, text2_lines)
                    print('\n'.join(diff))

                    kitap.save(file)
                    kitap.close()

                    mailGonder(textPD)

            def mailGonder(text):

                recipients = [sendMail]  # değişim varsa mail gönder
                emaillist = [elem.strip().split(',') for elem in recipients]

                msg = MIMEMultipart()
                msg['Subject'] = str(file)
                msg['From'] = 'usertolga@gmail.com'
                msg['Reply-to'] = sendMail
                msg.preamble = 'Multipart massage.\n'

                part = MIMEText(text)
                msg.attach(part)

                part = MIMEApplication(open(file, "rb").read())
                part.add_header('Content-Disposition', 'attachment', filename=file)
                msg.attach(part)

                server = smtplib.SMTP("smtp.gmail.com:587")
                server.ehlo()
                server.starttls()
                server.login("usertolga@gmail.com", "gmailsifresi")

                server.sendmail(msg['From'], emaillist, msg.as_string())

                print("Mail gönderilmiştir")



            try:  # kayıtlı data varsa çek
                kitap = openpyxl.load_workbook(file)
                sayfa = kitap.worksheets[0]

            except IOError as e:  # kayıtlı data yoksa oluştur
                kitap = Workbook()
                sayfa = kitap.active
                kitap.save(file)

            C1_verisi = sayfa['C1'].value  # başlangıç verisi
            Z1_verisi = sayfa['Z1'].value  # son veri yedeği

            if (C1_verisi == None or Z1_verisi == None):  # sütunlar boşsa başlangıç verilerini ekle
                sayfa['C1'] = sayfa['Z1'] = str(data)
                sayfa['A1'] = datetime.datetime.now()
                sayfa['I1'] = webLink
                textPI = "Veri dosyası oluşturulmuş ve ilk veri işlenmiştir!"
                print(data)
                print("Veri dosyası oluşturulmuş ve ilk veri işlenmiştir!")

                kitap.save(file)
                kitap.close()

                mailGonder(textPI)
            else:
                toplamSutun = sayfa.max_row

                for i in range(toplamSutun, 0, -1):
                    Ii = "I" + str(i)
                    Ci = "C" + str(i)
                    Ci_Verisi = sayfa[Ci].value
                    ##print((i, Ci_Verisi))
                    if (webLink == sayfa[Ii].value):
                        degisiklik(Ci_Verisi)
                        break
                    else:
                        toplamSutun = toplamSutun - 1
                        if (i == 1):
                            textPD = "Yeni URL adresine ait veri eklendi"
                            print("Yeni URL adresine ait veri eklendi")
                            sayfa.append({'C': data, 'A': datetime.datetime.now(), 'I': webLink})

                            kitap.save(file)
                            kitap.close()
                            mailGonder(textPD)
                            break

            print("\nGüncel data: \n" + data)
        time.sleep(20)


@app.route("/", methods=['GET', 'POST'])
def hello():
    form = ReusableForm(request.form)

    print
    form.errors
    if request.method == 'POST':
        name = request.form['name']
        mail = request.form['mail']

        calistir(name, mail)
        print
        name

        if form.validate():
            # Save the comment here.
            flash('Merhaba ' + name + ' nasılsın?' + mail)
        else:
            flash('Tüm formlar doldurulmalıdır! ')

    return render_template('hello.html', form=form)


if __name__ == "__main__":
    app.run()




